"""
Regression tests for spreadsheet formula injection (utils.to_excel): cells
beginning with =, +, -, or @ must be written as TEXT cells (data_type 's')
so Excel/LibreOffice render them literally instead of evaluating them.
"""

import io

import pandas as pd
from openpyxl import load_workbook

from utils import to_excel


def _first_column(df: pd.DataFrame) -> list:
    wb = load_workbook(io.BytesIO(to_excel(df)))
    ws = wb.active
    return [ws.cell(row=i, column=1) for i in range(2, ws.max_row + 1)]


def test_formula_like_strings_exported_as_text():
    df = pd.DataFrame({
        "description": ["=SUM(A1:A2)", "+1+1", "-2+3", "@cmd", "plain text"],
    })
    cells = _first_column(df)
    for cell in cells:
        assert cell.data_type == "s"  # text, never 'f' (formula)
    # The leading quote is stored and displayed — an inert text prefix.
    assert cells[0].value == "'=SUM(A1:A2)"
    assert cells[1].value == "'+1+1"
    assert cells[2].value == "'-2+3"
    assert cells[3].value == "'@cmd"
    assert cells[4].value == "plain text"


def test_non_string_cells_are_not_touched():
    df = pd.DataFrame({"x": [5, -7.5, True]})
    wb = load_workbook(io.BytesIO(to_excel(df)))
    ws = wb.active
    values = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    assert values == [5, -7.5, True]
